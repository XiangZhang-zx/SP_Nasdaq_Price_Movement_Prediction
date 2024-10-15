import os
import pandas as pd
import numpy as np
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
from tensorflow.keras.utils import to_categorical
from sklearn.preprocessing import LabelEncoder
import openpyxl
from openpyxl import load_workbook
from keras import backend as K
import gc
import tensorflow as tf
import time
# os.environ["CUDA_VISIBLE_DEVICES"] = "0"

# 设置文件路径
file_path_oex = '/usr3/graduate/xz0224/normal/oex.csv'
file_path_cash = '/usr3/graduate/xz0224/normal/cash.csv'

# 读取数据
data_oex = pd.read_csv(file_path_oex)
data_cash = pd.read_csv(file_path_cash)

# 确保日期列格式为datetime
data_oex['date'] = pd.to_datetime(data_oex['date'])
data_cash['date'] = pd.to_datetime(data_cash['date'])

# 创建 return_label 列
data_oex['return_label'] = data_oex['return'].apply(lambda x: '+' if x > 0 else '-')
data_cash['return_label'] = data_cash['return'].apply(lambda x: '+' if x > 0 else '-')

# 构建特征
def create_label_sequence(data):
    sequences = []
    for i in range(len(data)):
        if i >= 9:  # 确保有足够的数据来创建10天的序列
            sequence = ''.join(data['return_label'][i-10:i])
            sequences.append(sequence)
        else:
            sequences.append(None)  # 对于序列开始的部分，填充None
    return sequences

data_oex['feature'] = create_label_sequence(data_oex)
data_cash['feature'] = create_label_sequence(data_cash)

# 合并特征和目标变量，只保留oex的特征
combined_data = pd.merge(data_oex, data_cash, on='date')
combined_data['combined_feature'] = combined_data['feature_x']  # 只使用oex的特征
combined_data['target'] = combined_data.apply(lambda row: 'oex' if row['return_x'] > row['return_y'] else 'cash', axis=1)

# 设置时间范围
start_date = pd.to_datetime('2000-01-01')
end_date = pd.to_datetime('2024-06-30')

# 初始化投资金额
initial_investment = 100
current_value = initial_investment

# 加载已有的xlsx文件，如果不存在则创建一个新的
file_path = '/usr3/graduate/xz0224/normal/oex/investment_results_final(2006)(train_6month_c-c).xlsx'
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    # 找到最后一行的日期和投资金额
    last_row = sheet.max_row
    current_date = pd.to_datetime(sheet.cell(row=last_row, column=1).value) + pd.Timedelta(days=1)
    current_value = sheet.cell(row=last_row, column=2).value
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Date', 'Investment_Value', 'Decision'])
    current_date = start_date

# 准备用于绘图的数据
dates = [current_date]
values = [current_value]

# 循环遍历每一天
while current_date <= end_date:
    print(f"Starting iteration for date: {current_date} at time: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    # 设置训练集的时间范围（前五年）
    train_start = current_date - pd.DateOffset(months=6)
    train_end = current_date - pd.Timedelta(days=1)

    # 分割训练集和测试集
    train_set = combined_data[(combined_data['date'] >= train_start) & (combined_data['date'] <= train_end)]
    test_set = combined_data[combined_data['date'] == current_date]
    
    # 检查数据泄漏
    if not test_set.empty:
        assert not any(test_set.index.isin(train_set.index)), "Data leakage detected! Test set data found in the training set."
    
    # 在创建特征后，删除带有 None 值的行
    train_set = train_set.dropna(subset=['combined_feature'])
    test_set = test_set.dropna(subset=['combined_feature'])

    if not test_set.empty:
        # 现在进行 LSTM 模型训练的其他步骤
        X_train = train_set['combined_feature']
        y_train = train_set['target']
        X_test = test_set['combined_feature']
        y_test = test_set['target']

        # 将字符串序列转换为数值
        tokenizer = Tokenizer(char_level=True)
        tokenizer.fit_on_texts(X_train)
        X_train_seq = tokenizer.texts_to_sequences(X_train)
        X_test_seq = tokenizer.texts_to_sequences(X_test)

        # 确保所有序列具有相同的长度
        max_length = max([len(seq) for seq in X_train_seq])
        X_train_padded = pad_sequences(X_train_seq, maxlen=max_length, padding='post')
        X_test_padded = pad_sequences(X_test_seq, maxlen=max_length, padding='post')

        # LSTM 输入形状需要为 (样本数, 时间步长, 特征数)
        X_train_padded = np.expand_dims(X_train_padded, axis=-1)
        X_test_padded = np.expand_dims(X_test_padded, axis=-1)

        # 将目标变量转换为分类编码
        label_encoder = LabelEncoder()
        y_train_encoded = label_encoder.fit_transform(y_train)
        y_test_encoded = label_encoder.transform(y_test)
        y_train_categorical = to_categorical(y_train_encoded)
        y_test_categorical = to_categorical(y_test_encoded)

        # 创建 LSTM 模型
        model = Sequential()
        model.add(LSTM(50, input_shape=(max_length, 1)))
        model.add(Dense(2, activation='softmax'))  # 假设有两个输出类别

        # 编译模型
        model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

        # 训练模型
        model.fit(X_train_padded, y_train_categorical, epochs=10, batch_size=32)

        # 进行模型预测
        y_pred = model.predict(X_test_padded)

        # 将预测转换为具体的投资决策（'oex' 或 'cash'）
        decisions = np.argmax(y_pred, axis=1)

        # 模拟投资过程
        for i, decision in enumerate(decisions):
            # 获取对应日期的收益率
            date = test_set.iloc[i]['date']
            oex_rate = data_oex.loc[data_oex['date'] == date, 'return'].values
            cash_rate = data_cash.loc[data_cash['date'] == date, 'return'].values

            # 检查收益率是否为 None，如果是则设为 0
            oex_rate = oex_rate[0] if len(oex_rate) > 0 else 0
            cash_rate = cash_rate[0] if len(cash_rate) > 0 else 0

            # 根据预测决策更新投资金额
            if decision == 1:  # 'oex'
                current_value *= (1 + oex_rate)
            else:  # 'cash'
                current_value *= (1 + cash_rate)

            # 记录日期、金额和决策
            sheet.append([date, current_value, 'oex' if decision == 1 else 'cash'])
            wb.save(file_path)

        # 清理模型和内存
        del model
        del train_set, test_set, X_train, y_train, X_test, y_test
        del tokenizer, X_train_seq, X_test_seq, X_train_padded, X_test_padded
        del label_encoder, y_train_encoded, y_test_encoded, y_train_categorical, y_test_categorical
        K.clear_session()
        gc.collect()
        print(f"Investment amount: {current_value} Date: {current_date}")

        # 使用 TensorFlow 函数来释放 GPU 内存
    current_date += pd.Timedelta(days=1)

# 保存最终的结果
wb.save(file_path)
